# -*- coding: UTF-8 -*-
from xml.dom import minidom
import xlrd
import openpyxl
import requests
import json
import sys
import HTMLParser
import os
import re
import codecs
import time
import datetime

class OptionExcelData(object):
     #对Excel进行操作，包括读取请求参数，和填写操作结果
     def __init__(self, excelFile, excelPath=''):
          self.excelFile = excelFile
          self.excelPath = excelPath
          self.caseList = []

def getCaseList(self,excelFile,excelPath=''):
     readExcel = xlrd.open_workbook(fileName) #读取指定的Excel
     try:
          table = readExcel.sheet_by_index(0)  #获取Excel的第一个sheet
          trows = table.nrows  #获取Excel的行数
          for n in range(1, trows):
               tmpdict = {}   #把一行记录写进一个{}
               tmpdict['id'] = n  #n是Excel中的第n行
               tmpdict['casename'] = table.cell(n, 2).value
               tmpdict['method'] = table.cell(n, 3).value
               tmpdict['request'] = table.cell(n, 4).value
               self.caseList.append(tmpdict)
     except Exception, e:
          raise
     finally:
          pass
     return self.caseList